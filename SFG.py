%pip install openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# Load the data from Excel files
file3 = pd.read_excel("file3.xlsx")
file4 = pd.read_excel("file4.xlsx")

# Add a column for Physical Stock
file3['Physical Stock'] = 0

# Iterate through each row in file4
for index, row in file4.iterrows():
    # Check if the material code exists in file3
    if row['Material'] in file3['Material'].values:
        # Find the index of the material code in file3
        idx = file3.index[file3['Material'] == row['Material']].tolist()[0]
        # Create a new row for the new batch
        new_row = row.copy()
        new_row['Unrestricted'] = 0
        new_row['In Quality Insp.'] = 0
        new_row['Blocked'] = 0
        new_row['Physical Stock'] = row['Physical Stock']
        # Append the new row below the existing row
        file3 = pd.concat([file3.iloc[:idx+1], pd.DataFrame([new_row]), file3.iloc[idx+1:]], ignore_index=True)
    else:
        # Create a new row with "NA" for missing columns
        new_row = {column: 'NA' for column in file3.columns}
        new_row.update(row)  # Update with data from file 4
        # Append the new row to file3
        file3 = pd.concat([file3, pd.DataFrame([new_row])], ignore_index=True)

# Reindex the dataframe
file3.reset_index(drop=True, inplace=True)

# Create a new Excel writer object
writer = pd.ExcelWriter("output.xlsx", engine='openpyxl')
file3.to_excel(writer, index=False, sheet_name='Sheet1')

# Get the workbook and worksheet objects
workbook  = writer.book
worksheet = writer.sheets['Sheet1']

# Add borders to all cells
border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

for row in worksheet.iter_rows():
    for cell in row:
        cell.border = border

# Wrap text and apply grey background to headers
grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
for cell in worksheet[1]:
    cell.fill = grey_fill
    cell.alignment = Alignment(wrap_text=True)

# Apply cell formatting
for column in worksheet.columns:
    max_length = 0
    column = [cell for cell in column if cell.value is not None]
    max_length = max((len(str(cell.value)) for cell in column))
    adjusted_width = (max_length + 2) * 1.2
    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    for cell in column:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=10)
        # Custom number formatting for specific columns
        if cell.column_letter == 'H':  # Batch column
            cell.number_format = '000'
        elif cell.column_letter == 'I':  # Physical Stock column
            cell.number_format = '0.000'
        elif cell.column_letter == 'A':  # Material column
            cell.number_format = '0000000000'

# Close the writer
writer.close()
